from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import time
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Callable, Iterable
from urllib.parse import parse_qs, quote_plus, urljoin, urlsplit, urlunsplit

import requests
from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


BASE_TAG_URL = "https://vietnamnet.vn/lai-suat-tag13172411016833496207.html"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/126.0 Safari/537.36"
)
STANDARD_MONTHS = [1, 3, 6, 9, 12, 18]
ProgressCallback = Callable[[str, dict[str, object]], None]
VIETNAMNET_START_DATE = date(2023, 5, 31)
LEGACY_DOMAINS = [
    "thanhnien.vn",
    "thanhtra.com.vn",
    "vneconomy.vn",
    "thesaigontimes.vn",
    "tinnhanhchungkhoan.vn",
    "vietnamfinance.vn",
    "baodautu.vn",
]
BANK_ALIASES = {
    "AGRIBANK": ["agribank", "ngan hang nong nghiep"],
    "BIDV": ["bidv"],
    "VIETCOMBANK": ["vietcombank", "vcb"],
    "VIETINBANK": ["vietinbank"],
    "ACB": ["acb"],
    "MBBANK": ["mbbank", "mb bank", "ngan hang quan doi"],
    "SACOMBANK": ["sacombank"],
    "TECHCOMBANK": ["techcombank"],
    "VPBANK": ["vpbank"],
    "EXIMBANK": ["eximbank"],
    "HDBANK": ["hdbank"],
    "SHB": ["shb"],
    "VIB": ["vib"],
    "OCB": ["ocb"],
    "MSB": ["msb"],
    "SCB": ["scb"],
    "SEABANK": ["seabank"],
    "NAM A BANK": ["nam a bank", "namabank"],
    "NCB": ["ncb"],
    "PVCOMBANK": ["pvcombank"],
    "BAC A BANK": ["bac a bank", "baca bank", "bacabank"],
    "PGBANK": ["pgbank"],
    "VIET A BANK": ["viet a bank", "vietabank"],
    "VIETBANK": ["vietbank"],
    "KIENLONGBANK": ["kienlongbank"],
    "BAOVIETBANK": ["baovietbank", "bao viet bank"],
}


@dataclass(frozen=True)
class ArticleLink:
    title: str
    url: str
    found_on_page: int


@dataclass
class RateRecord:
    article_date: date | None
    bank: str
    rates: dict[int, float | None]
    source_title: str
    source_url: str
    table_title: str
    row_order: int


@dataclass
class ArticleLog:
    article_date: date | None
    title: str
    url: str
    status: str
    rows: int = 0
    error: str = ""
    found_on_page: int | None = None


def clean_text(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", " ", value).strip()


def fold_text(value: str | None) -> str:
    text = clean_text(value).replace("đ", "d").replace("Đ", "D")
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return text.lower()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Scrape Vietnamnet daily bank interest-rate tables to Excel."
    )
    parser.add_argument(
        "--menu",
        action="store_true",
        help="Open a simple interactive menu before scraping.",
    )
    parser.add_argument("--pages", type=int, default=500, help="Max tag pages to scan.")
    parser.add_argument("--start-page", type=int, default=0, help="First tag page number.")
    parser.add_argument("--base-url", default=BASE_TAG_URL, help="Vietnamnet tag URL.")
    parser.add_argument(
        "--output",
        default="outputs/lstk_vietnamnet.xlsx",
        help="Output .xlsx path.",
    )
    parser.add_argument(
        "--cache-dir",
        default=".cache/vietnamnet",
        help="HTML cache directory.",
    )
    parser.add_argument(
        "--refresh",
        action="store_true",
        help="Ignore cached HTML and fetch pages again.",
    )
    parser.add_argument("--delay", type=float, default=0.25, help="Delay between requests.")
    parser.add_argument("--timeout", type=float, default=25, help="Request timeout seconds.")
    parser.add_argument(
        "--max-articles",
        type=int,
        default=0,
        help="Optional cap for article pages. 0 means no cap.",
    )
    parser.add_argument(
        "--stop-after-empty",
        type=int,
        default=8,
        help="Stop after this many tag pages with no matching daily articles. 0 disables.",
    )
    parser.add_argument(
        "--date-from",
        default="",
        help="Optional inclusive start date, ISO yyyy-mm-dd.",
    )
    parser.add_argument(
        "--date-to",
        default="",
        help="Optional inclusive end date, ISO yyyy-mm-dd.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Print every skipped article/error.",
    )
    parser.add_argument(
        "--quiet",
        action="store_true",
        help="Suppress per-page and per-article progress output.",
    )
    parser.add_argument(
        "--legacy",
        action="store_true",
        help="Also search older free-text news sources before Vietnamnet table coverage.",
    )
    parser.add_argument(
        "--legacy-from-year",
        type=int,
        default=2015,
        help="First year for legacy source search.",
    )
    parser.add_argument(
        "--legacy-to-year",
        type=int,
        default=2023,
        help="Last year for legacy source search.",
    )
    parser.add_argument(
        "--legacy-period",
        choices=("month", "day"),
        default="month",
        help="Search legacy sources by month or by day.",
    )
    parser.add_argument(
        "--legacy-max-results",
        type=int,
        default=10,
        help="Max Google results to inspect per legacy period.",
    )
    parser.add_argument(
        "--legacy-max-articles",
        type=int,
        default=0,
        help="Optional cap for fetched legacy articles. 0 means no cap.",
    )
    parser.add_argument(
        "--legacy-only",
        action="store_true",
        help="Skip Vietnamnet archive and only run legacy source search.",
    )
    return parser.parse_args()


def ask_int(prompt: str, default: int, minimum: int = 0) -> int:
    while True:
        try:
            raw = input(f"{prompt} [{default}]: ").strip()
        except EOFError:
            return default
        if not raw:
            return default
        try:
            value = int(raw)
        except ValueError:
            print("Please enter a whole number.")
            continue
        if value < minimum:
            print(f"Please enter {minimum} or higher.")
            continue
        return value


def ask_text(prompt: str, default: str) -> str:
    try:
        raw = input(f"{prompt} [{default}]: ").strip()
    except EOFError:
        return default
    return raw or default


def ask_yes_no(prompt: str, default: bool = False) -> bool:
    suffix = "Y/n" if default else "y/N"
    while True:
        try:
            raw = input(f"{prompt} [{suffix}]: ").strip().lower()
        except EOFError:
            return default
        if not raw:
            return default
        if raw in {"y", "yes"}:
            return True
        if raw in {"n", "no"}:
            return False
        print("Please enter y or n.")


def apply_menu(args: argparse.Namespace) -> argparse.Namespace:
    title = "PASSION INVESTMENT TOOL LAY LSTK"
    border = "=" * len(title)
    print()
    print(border)
    print(title)
    print(border)
    print("Simple launcher for scraping Vietnamnet bank-rate tables into Excel.")
    print()

    args.pages = ask_int("How many tag pages should I scan?", args.pages, minimum=1)
    args.output = ask_text("Output Excel file", args.output)
    args.refresh = ask_yes_no("Refresh downloaded pages instead of using cache?", args.refresh)
    args.stop_after_empty = ask_int(
        "Stop after how many empty archive pages?", args.stop_after_empty, minimum=0
    )
    args.quiet = False

    print("Ready to run:")
    print(f"  Pages to scan : {args.pages}")
    print(f"  Output file   : {args.output}")
    print(f"  Refresh cache : {'yes' if args.refresh else 'no'}")
    print(f"  Empty-page stop: {args.stop_after_empty}")
    print()
    print("Starting scrape...")
    print()
    return args


def progress_line(done: int, total: int, label: str) -> str:
    if total <= 0:
        return f"{label}: {done}"
    pct = (done / total) * 100
    return f"{label}: {done}/{total} ({pct:5.1f}%)"


def emit_progress(
    progress_callback: ProgressCallback | None,
    event: str,
    **payload: object,
) -> None:
    if progress_callback is not None:
        progress_callback(event, payload)


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})
    retry = Retry(
        total=3,
        read=3,
        connect=3,
        status=3,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET",),
        backoff_factor=0.6,
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    return session


def cache_path(cache_dir: Path, url: str) -> Path:
    digest = hashlib.sha1(url.encode("utf-8")).hexdigest()
    return cache_dir / f"{digest}.html"


def fetch_html(
    session: requests.Session,
    url: str,
    cache_dir: Path,
    timeout: float,
    refresh: bool,
    delay: float,
) -> str:
    cache_dir.mkdir(parents=True, exist_ok=True)
    path = cache_path(cache_dir, url)
    if path.exists() and not refresh:
        return path.read_text(encoding="utf-8", errors="replace")

    response = session.get(url, timeout=timeout)
    response.raise_for_status()
    response.encoding = response.encoding or "utf-8"
    html = response.text
    path.write_text(html, encoding="utf-8")
    if delay > 0:
        time.sleep(delay)
    return html


def index_url(base_url: str, page: int) -> str:
    if page <= 0:
        return base_url
    if base_url.endswith(".html"):
        return base_url[:-5] + f"-page{page}.html"
    return base_url.rstrip("/") + f"-page{page}.html"


def canonical_url(url: str, base_url: str = BASE_TAG_URL) -> str:
    absolute = urljoin(base_url, url)
    split = urlsplit(absolute)
    return urlunsplit((split.scheme, split.netloc, split.path, "", ""))


def is_daily_rate_article(title: str) -> bool:
    return "lai suat ngan hang hom nay" in fold_text(title)


def build_date(year: int, month: int, day: int) -> date | None:
    max_reasonable_year = date.today().year + 2
    if year < 2000 or year > max_reasonable_year:
        return None
    try:
        return date(year, month, day)
    except ValueError:
        return None


def parse_date_from_text(*values: str, default_year: int | None = None) -> date | None:
    for value in values:
        text = clean_text(value)
        for day_s, month_s, year_s in re.findall(
            r"(?<!\d)(\d{1,2})[/-](\d{1,2})[/-](20\d{2})(?!\d)", text
        ):
            parsed = build_date(int(year_s), int(month_s), int(day_s))
            if parsed:
                return parsed

    if default_year is None:
        return None

    for value in values:
        text = clean_text(value)
        for day_s, month_s in re.findall(
            r"(?<!\d)(\d{1,2})/(\d{1,2})(?!\s*/?\s*\d)", text
        ):
            parsed = build_date(default_year, int(month_s), int(day_s))
            if parsed:
                return parsed
    return None


def parse_date_from_url(url: str) -> date | None:
    path = urlsplit(url).path
    for day_s, month_s, year_s in re.findall(
        r"(?<!\d)(\d{1,2})-(\d{1,2})-(20\d{2})(?!\d)", path
    ):
        parsed = build_date(int(year_s), int(month_s), int(day_s))
        if parsed:
            return parsed
    return None


def parse_iso_like_date(value: str | None) -> date | None:
    if not value:
        return None
    match = re.search(r"(?<!\d)(20\d{2})-(\d{1,2})-(\d{1,2})(?!\d)", value)
    if not match:
        return None
    year_s, month_s, day_s = match.groups()
    return build_date(int(year_s), int(month_s), int(day_s))


def parse_published_date(soup: BeautifulSoup) -> date | None:
    meta_selectors = [
        "meta[property='article:published_time']",
        "meta[property='og:published_time']",
        "meta[name='pubdate']",
        "meta[itemprop='datePublished']",
    ]
    for selector in meta_selectors:
        meta = soup.select_one(selector)
        parsed = parse_iso_like_date(meta.get("content") if meta else None)
        if parsed:
            return parsed

    for script in soup.find_all("script"):
        text = script.string or script.get_text(" ", strip=False)
        if not text or ("datePublished" not in text and "ArticlePublishDate" not in text):
            continue
        parsed = parse_iso_like_date(text)
        if parsed:
            return parsed
    return None


def vietnamese_score(value: str) -> int:
    markers = "ăâđêôơưáàảãạấầẩẫậắằẳẵặéèẻẽẹếềểễệíìỉĩịóòỏõọốồổỗộớờởỡợúùủũụứừửữựýỳỷỹỵ"
    return sum(1 for char in value.lower() if char in markers)


def repair_mojibake(value: str) -> str:
    try:
        repaired = value.encode("cp1252", errors="ignore").decode("utf-8", errors="ignore")
    except UnicodeError:
        return value
    return repaired if vietnamese_score(repaired) > vietnamese_score(value) else value


def visible_text_from_html(html: str) -> tuple[str, str, date | None]:
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "noscript", "svg"]):
        tag.decompose()
    title = clean_text(
        soup.select_one("h1").get_text(" ", strip=True)
        if soup.select_one("h1")
        else (soup.title.get_text(" ", strip=True) if soup.title else "")
    )
    text = repair_mojibake(soup.get_text("\n", strip=True))
    title = repair_mojibake(title)
    return title, text, parse_published_date(soup)


def allowed_legacy_url(url: str) -> bool:
    host = urlsplit(url).netloc.lower().removeprefix("www.")
    return any(host == domain or host.endswith(f".{domain}") for domain in LEGACY_DOMAINS)


def google_result_url(href: str) -> str:
    if href.startswith("/url?"):
        return parse_qs(urlsplit(href).query).get("q", [""])[0]
    if href.startswith("http"):
        return href
    return ""


def legacy_period_ranges(start: date, end: date, period: str) -> Iterable[tuple[date, date]]:
    current = start
    while current <= end:
        if period == "day":
            next_start = current + timedelta(days=1)
        else:
            if current.month == 12:
                next_start = date(current.year + 1, 1, 1)
            else:
                next_start = date(current.year, current.month + 1, 1)
        yield current, min(next_start, end + timedelta(days=1))
        current = next_start


def google_search_links(
    session: requests.Session,
    args: argparse.Namespace,
    period_start: date,
    period_end_exclusive: date,
) -> list[str]:
    site_query = " OR ".join(f"site:{domain}" for domain in LEGACY_DOMAINS)
    query = (
        f"({site_query}) \"lãi suất\" "
        f"after:{period_start:%Y-%m-%d} before:{period_end_exclusive:%Y-%m-%d}"
    )
    url = f"https://www.google.com/search?q={quote_plus(query)}&num={args.legacy_max_results}&hl=vi"
    try:
        html = fetch_html(
            session,
            url,
            Path(args.cache_dir) / "legacy_google",
            args.timeout,
            args.refresh,
            args.delay,
        )
    except Exception:
        return []

    soup = BeautifulSoup(html, "lxml")
    links: list[str] = []
    seen: set[str] = set()
    for anchor in soup.select("a[href]"):
        raw_link = google_result_url(anchor.get("href", ""))
        if not raw_link:
            continue
        link = canonical_url(raw_link)
        if not link or not allowed_legacy_url(link) or link in seen:
            continue
        links.append(link)
        seen.add(link)
        if len(links) >= args.legacy_max_results:
            break
    return links


def banks_in_text(value: str) -> list[str]:
    folded = fold_text(value)
    banks: list[str] = []
    for bank, aliases in BANK_ALIASES.items():
        if any(alias in folded for alias in aliases):
            banks.append(bank)
    return banks


def split_legacy_blocks(text: str) -> list[str]:
    blocks = [clean_text(block) for block in re.split(r"[\n\r]+|(?<=[.!?])\s+", text)]
    return [block for block in blocks if block]


def is_vnd_savings_context(value: str) -> bool:
    folded = fold_text(value)
    if any(token in folded for token in ("usd", "eur", "ngoai te", "dollar")):
        return False
    return any(token in folded for token in ("vnd", "tiet kiem", "huy dong", "tien gui", "ky han"))


def extract_term_rates(block: str) -> dict[int, float]:
    rates: dict[int, float] = {}
    folded = fold_text(block)
    term_pattern = re.compile(
        r"(?:tu\s*)?(\d{1,2})\s*(?:[-–]\s*(\d{1,2}))?\s*thang"
    )
    matches = list(term_pattern.finditer(folded))

    def standard_terms(start_term: int, end_term: int | None) -> list[int]:
        if end_term is None:
            return [start_term] if start_term in STANDARD_MONTHS else []
        low, high = sorted((start_term, end_term))
        return [term for term in STANDARD_MONTHS if low <= term <= high]

    def current_rate(segment: str) -> float | None:
        preferred = re.search(
            r"(?:xuong\s+con|giam\s+con|con|la|chi)\s*(\d{1,2}(?:[,.]\d{1,2})?)\s*%",
            segment,
        )
        if preferred:
            return parse_rate(preferred.group(1))
        candidates = re.findall(r"(\d{1,2}(?:[,.]\d{1,2})?)\s*%", segment)
        if not candidates:
            return None
        return parse_rate(candidates[-1] if "xuong" in segment or "giam" in segment else candidates[0])

    for index, match in enumerate(matches):
        start_term = int(match.group(1))
        end_term = int(match.group(2)) if match.group(2) else None
        terms = standard_terms(start_term, end_term)
        if not terms:
            continue
        next_start = matches[index + 1].start() if index + 1 < len(matches) else len(folded)
        segment = folded[match.end() : min(next_start, match.end() + 180)]
        rate = current_rate(segment)
        if rate is None or rate > 20:
            continue
        for term in terms:
            rates[term] = rate
    return rates


def extract_legacy_records(
    title: str,
    text: str,
    source_url: str,
    article_date: date | None,
) -> list[RateRecord]:
    detected_from_title = banks_in_text(title)
    active_bank = detected_from_title[0] if len(detected_from_title) == 1 else None
    bank_rates: dict[str, dict[int, float | None]] = {}

    for block in split_legacy_blocks(text):
        mentioned = banks_in_text(block)
        rates = extract_term_rates(block) if is_vnd_savings_context(block) else {}
        if mentioned and rates:
            for bank in mentioned:
                bank_rates.setdefault(bank, {})
                bank_rates[bank].update(rates)
            active_bank = mentioned[0] if len(mentioned) == 1 else active_bank
            continue
        if mentioned:
            active_bank = mentioned[0] if len(mentioned) == 1 else None
        if not active_bank or not is_vnd_savings_context(block):
            continue
        if not rates:
            continue
        bank_rates.setdefault(active_bank, {})
        bank_rates[active_bank].update(rates)

    records: list[RateRecord] = []
    for row_order, (bank, rates) in enumerate(bank_rates.items(), start=1):
        records.append(
            RateRecord(
                article_date=article_date,
                bank=bank,
                rates=rates,
                source_title=title,
                source_url=source_url,
                table_title="Legacy text extraction",
                row_order=row_order,
            )
        )
    return records


def scrape_legacy_article(
    session: requests.Session,
    url: str,
    args: argparse.Namespace,
) -> tuple[list[RateRecord], ArticleLog]:
    html = fetch_html(
        session,
        url,
        Path(args.cache_dir) / "legacy_articles",
        args.timeout,
        args.refresh,
        args.delay,
    )
    title, text, published_date = visible_text_from_html(html)
    article_date = (
        parse_date_from_text(title, default_year=published_date.year if published_date else None)
        or parse_date_from_url(url)
        or published_date
    )
    records = extract_legacy_records(title, text, url, article_date)
    status = "ok" if records else "no legacy rates"
    return records, ArticleLog(
        article_date=article_date,
        title=title or url,
        url=url,
        status=status,
        rows=len(records),
        error="" if records else "No VND bank/term/rate pairs were detected.",
    )


def run_legacy_scrape(
    session: requests.Session,
    args: argparse.Namespace,
    progress_callback: ProgressCallback | None = None,
) -> tuple[list[RateRecord], list[ArticleLog]]:
    start = date(args.legacy_from_year, 1, 1)
    end = min(date(args.legacy_to_year, 12, 31), VIETNAMNET_START_DATE - timedelta(days=1))
    if start > end:
        return [], []

    periods = list(legacy_period_ranges(start, end, args.legacy_period))
    records: list[RateRecord] = []
    logs: list[ArticleLog] = []
    seen_urls: set[str] = set()
    fetched = 0

    for idx, (period_start, period_end) in enumerate(periods, start=1):
        links = google_search_links(session, args, period_start, period_end)
        emit_progress(
            progress_callback,
            "legacy_period",
            done=idx,
            total=len(periods),
            period=f"{period_start:%Y-%m-%d} to {(period_end - timedelta(days=1)):%Y-%m-%d}",
            links=len(links),
        )
        for url in links:
            if url in seen_urls:
                continue
            if args.legacy_max_articles and fetched >= args.legacy_max_articles:
                return records, logs
            seen_urls.add(url)
            fetched += 1
            try:
                article_records, log = scrape_legacy_article(session, url, args)
            except Exception as exc:  # noqa: BLE001 - keep legacy search moving.
                log = ArticleLog(None, url, url, "error", error=str(exc))
                article_records = []
            records.extend(article_records)
            logs.append(log)
            emit_progress(
                progress_callback,
                "legacy_article",
                done=fetched,
                total=max(fetched, args.legacy_max_articles or fetched),
                status=log.status,
                rows=log.rows,
                title=log.title,
                url=log.url,
            )

    return records, logs


def parse_iso_date(value: str) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as exc:
        raise SystemExit(f"Invalid date '{value}'. Use yyyy-mm-dd.") from exc


def date_in_range(value: date | None, date_from: date | None, date_to: date | None) -> bool:
    if value is None:
        return True
    if date_from and value < date_from:
        return False
    if date_to and value > date_to:
        return False
    return True


def collect_article_links(
    session: requests.Session,
    args: argparse.Namespace,
    date_from: date | None,
    date_to: date | None,
    progress_callback: ProgressCallback | None = None,
) -> list[ArticleLink]:
    cache_dir = Path(args.cache_dir)
    seen: set[str] = set()
    links: list[ArticleLink] = []
    empty_streak = 0
    last_page = args.start_page + max(args.pages, 0)

    for page in range(args.start_page, last_page):
        url = index_url(args.base_url, page)
        html = fetch_html(session, url, cache_dir, args.timeout, args.refresh, args.delay)
        soup = BeautifulSoup(html, "lxml")
        page_matches = 0
        page_new = 0

        for anchor in soup.select("a[href]"):
            title = clean_text(anchor.get_text(" ", strip=True))
            if not title or not is_daily_rate_article(title):
                continue

            article_url = canonical_url(anchor.get("href", ""), args.base_url)
            if not article_url.endswith(".html"):
                continue
            if "vietnamnet.vn" not in urlsplit(article_url).netloc:
                continue

            page_matches += 1
            if article_url in seen:
                continue

            article_date = parse_date_from_text(title) or parse_date_from_url(article_url)
            if not date_in_range(article_date, date_from, date_to):
                seen.add(article_url)
                continue

            links.append(ArticleLink(title=title, url=article_url, found_on_page=page))
            seen.add(article_url)
            page_new += 1
            if args.max_articles and len(links) >= args.max_articles:
                if not args.quiet:
                    print(f"Collected {len(links)} article links; hit --max-articles.")
                emit_progress(
                    progress_callback,
                    "page_scan",
                    page=page,
                    pages_done=page - args.start_page + 1,
                    pages_total=args.pages,
                    matches=page_matches,
                    new_links=page_new,
                    total_links=len(links),
                )
                return links

        if page_matches:
            empty_streak = 0
        else:
            empty_streak += 1

        if not args.quiet:
            print(
                f"Tag page {page}: {page_matches} matching links, "
                f"{page_new} new, {len(links)} total. "
                f"{progress_line(page - args.start_page + 1, args.pages, 'Page scan')}"
            )
        emit_progress(
            progress_callback,
            "page_scan",
            page=page,
            pages_done=page - args.start_page + 1,
            pages_total=args.pages,
            matches=page_matches,
            new_links=page_new,
            total_links=len(links),
        )

        if args.stop_after_empty and empty_streak >= args.stop_after_empty:
            if not args.quiet:
                print(
                    f"Stopping after {empty_streak} empty tag pages "
                    f"(last checked page {page})."
                )
            emit_progress(
                progress_callback,
                "page_scan_done",
                pages_done=page - args.start_page + 1,
                pages_total=args.pages,
                total_links=len(links),
            )
            break

    return links


def direct_cells(row: Tag) -> list[Tag]:
    return [cell for cell in row.find_all(["td", "th"], recursive=False)]


def parse_span(value: str | None) -> int:
    try:
        parsed = int(value or "1")
    except ValueError:
        return 1
    return max(parsed, 1)


def expand_html_table(table: Tag) -> list[list[str]]:
    rows: list[list[str]] = []
    spans: dict[int, list[object]] = {}

    for tr in table.find_all("tr"):
        output: list[str] = []
        col = 0

        def fill_spans() -> None:
            nonlocal col
            while col in spans:
                text, remaining = spans[col]
                output.append(str(text))
                remaining = int(remaining) - 1
                if remaining <= 0:
                    del spans[col]
                else:
                    spans[col] = [text, remaining]
                col += 1

        for cell in direct_cells(tr):
            fill_spans()
            text = clean_text(cell.get_text(" ", strip=True))
            colspan = parse_span(cell.get("colspan"))
            rowspan = parse_span(cell.get("rowspan"))

            for offset in range(colspan):
                output.append(text)
                if rowspan > 1:
                    spans[col + offset] = [text, rowspan - 1]
            col += colspan

        fill_spans()
        if any(value for value in output):
            rows.append(output)

    return rows


def locate_header_row(rows: list[list[str]]) -> int | None:
    for idx, row in enumerate(rows):
        folded = [fold_text(cell) for cell in row]
        has_bank = any("ngan hang" in cell for cell in folded)
        month_count = sum(1 for cell in folded if "thang" in cell)
        if has_bank and month_count >= 3:
            return idx
    return None


def month_from_header(header: str) -> int | None:
    match = re.search(r"(\d{1,2})\s*thang", fold_text(header))
    if not match:
        return None
    return int(match.group(1))


def parse_rate(value: str) -> float | None:
    text = clean_text(value)
    if not text or fold_text(text) in {"-", "na", "n/a", "khong"}:
        return None
    text = text.replace("%", "").replace("\u2212", "-")
    match = re.search(r"-?\d+(?:[,.]\d+)?", text)
    if not match:
        return None
    return float(match.group(0).replace(",", "."))


def is_bank_row(row: list[str], header_len: int) -> bool:
    if not row:
        return False
    bank = clean_text(row[0])
    if not bank:
        return False
    folded = fold_text(bank)
    if "ngan hang" in folded or "lai suat" in folded or "bieu" in folded:
        return False
    return len(row) >= min(2, header_len)


def extract_rate_table(
    soup: BeautifulSoup,
    article_date: date | None,
    source_title: str,
    source_url: str,
) -> tuple[list[RateRecord], str]:
    best_records: list[RateRecord] = []
    best_title = ""

    for table in soup.select("table"):
        rows = expand_html_table(table)
        header_idx = locate_header_row(rows)
        if header_idx is None:
            continue

        header = rows[header_idx]
        month_columns: dict[int, int] = {}
        for col_idx, header_text in enumerate(header):
            month = month_from_header(header_text)
            if month is not None:
                month_columns[month] = col_idx

        if len(month_columns) < 3:
            continue

        caption = clean_text(table.caption.get_text(" ", strip=True)) if table.caption else ""
        title_parts = [caption] if caption else []
        for prior in rows[:header_idx]:
            unique_cells = list(dict.fromkeys(cell for cell in prior if cell))
            if len(unique_cells) <= 2:
                joined = clean_text(" ".join(unique_cells))
                if joined:
                    title_parts.append(joined)
        table_title = clean_text(" | ".join(dict.fromkeys(title_parts)))

        records: list[RateRecord] = []
        for row_order, row in enumerate(rows[header_idx + 1 :], start=1):
            if not is_bank_row(row, len(header)):
                continue
            bank = clean_text(row[0]).upper()
            rates: dict[int, float | None] = {}
            for month, col_idx in month_columns.items():
                raw = row[col_idx] if col_idx < len(row) else ""
                rates[month] = parse_rate(raw)

            if not any(value is not None for value in rates.values()):
                continue

            records.append(
                RateRecord(
                    article_date=article_date,
                    bank=bank,
                    rates=rates,
                    source_title=source_title,
                    source_url=source_url,
                    table_title=table_title,
                    row_order=row_order,
                )
            )

        if len(records) > len(best_records):
            best_records = records
            best_title = table_title

    return best_records, best_title


def article_title_and_url(soup: BeautifulSoup, fallback: ArticleLink) -> tuple[str, str]:
    h1 = soup.select_one("h1")
    title = clean_text(h1.get_text(" ", strip=True)) if h1 else fallback.title
    canonical = soup.select_one("link[rel='canonical']")
    url = fallback.url
    if canonical and canonical.get("href"):
        url = canonical_url(canonical["href"], fallback.url)
    return title, url


def scrape_article(
    session: requests.Session,
    link: ArticleLink,
    args: argparse.Namespace,
) -> tuple[list[RateRecord], ArticleLog]:
    html = fetch_html(
        session,
        link.url,
        Path(args.cache_dir),
        args.timeout,
        args.refresh,
        args.delay,
    )
    soup = BeautifulSoup(html, "lxml")
    title, source_url = article_title_and_url(soup, link)
    published_date = parse_published_date(soup)
    default_year = published_date.year if published_date else None
    article_date = (
        parse_date_from_text(title, default_year=default_year)
        or parse_date_from_url(source_url)
    )
    records, table_title = extract_rate_table(soup, article_date, title, source_url)

    if article_date is None:
        article_date = (
            parse_date_from_text(table_title, title, default_year=default_year)
            or parse_date_from_url(source_url)
            or published_date
        )

    for record in records:
        record.article_date = article_date

    if not records:
        return records, ArticleLog(
            article_date=article_date,
            title=title,
            url=source_url,
            status="no table",
            rows=0,
            error="No bank-rate table with NGAN HANG and month columns was found.",
            found_on_page=link.found_on_page,
        )

    return records, ArticleLog(
        article_date=article_date,
        title=title,
        url=source_url,
        status="ok",
        rows=len(records),
        found_on_page=link.found_on_page,
    )


def all_months(records: Iterable[RateRecord]) -> list[int]:
    found = {month for record in records for month in record.rates}
    ordered = [month for month in STANDARD_MONTHS if month in found]
    ordered.extend(sorted(month for month in found if month not in set(ordered)))
    return ordered


def safe_sheet_title(value: str) -> str:
    return re.sub(r"[\[\]:*?/\\]", "_", value)[:31] or "Sheet"


def add_table(ws, name: str) -> None:
    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def fit_columns(ws, max_width: int = 70) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = 10
        for row_idx in range(1, min(ws.max_row, 500) + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is None:
                continue
            width = max(width, min(len(str(value)) + 2, max_width))
        ws.column_dimensions[letter].width = width


def apply_common_sheet_style(ws) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    side = Side(style="thin", color="D9E2EC")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = Border(bottom=side)
    fit_columns(ws)


def write_workbook(
    records: list[RateRecord],
    article_logs: list[ArticleLog],
    output_path: Path,
    run_info: dict[str, object],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    months = STANDARD_MONTHS

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    ok_logs = [log for log in article_logs if log.status == "ok"]
    dates = sorted({record.article_date for record in records if record.article_date})
    banks = sorted({record.bank for record in records})
    summary_rows = [
        ("Generated at", datetime.now().replace(microsecond=0)),
        ("Source tag", run_info.get("base_url", BASE_TAG_URL)),
        ("Tag pages requested", run_info.get("pages", "")),
        ("Article links collected", len(article_logs)),
        ("Articles with tables", len(ok_logs)),
        ("Rate rows exported", len(records)),
        ("Unique article dates", len(dates)),
        ("Earliest date", dates[0] if dates else None),
        ("Latest date", dates[-1] if dates else None),
        ("Unique banks", len(banks)),
    ]
    summary.append(["Metric", "Value"])
    for row in summary_rows:
        summary.append(list(row))
    style_header(summary)
    summary["B2"].number_format = "yyyy-mm-dd hh:mm"
    for cell in ("B9", "B10"):
        summary[cell].number_format = "yyyy-mm-dd"
    summary["B3"].hyperlink = str(run_info.get("base_url", BASE_TAG_URL))
    summary["B3"].style = "Hyperlink"
    apply_common_sheet_style(summary)
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 70

    rates_ws = wb.create_sheet("All Rates")
    rate_headers = (
        ["date", "bank"]
        + [f"{month}_month_pct" for month in months]
        + ["source_title", "source_url", "table_title"]
    )
    rates_ws.append(rate_headers)
    sorted_records = sorted(
        records,
        key=lambda item: (
            -(item.article_date or date.min).toordinal(),
            item.row_order,
            item.bank,
            item.source_url,
        ),
    )
    for record in sorted_records:
        rates_ws.append(
            [record.article_date, record.bank]
            + [record.rates.get(month) for month in months]
            + [record.source_title, record.source_url, record.table_title]
        )
    style_header(rates_ws)
    for row in range(2, rates_ws.max_row + 1):
        rates_ws.cell(row, 1).number_format = "yyyy-mm-dd"
        for col in range(3, 3 + len(months)):
            rates_ws.cell(row, col).number_format = "0.00"
            rates_ws.cell(row, col).alignment = Alignment(horizontal="right")
        url_cell = rates_ws.cell(row, 4 + len(months))
        if url_cell.value:
            url_cell.hyperlink = str(url_cell.value)
            url_cell.style = "Hyperlink"
    add_table(rates_ws, "AllRates")
    apply_common_sheet_style(rates_ws)
    rates_ws.freeze_panes = "C2"

    latest_ws = wb.create_sheet("Latest Table")
    if sorted_records:
        latest_date = max(record.article_date for record in records if record.article_date)
        latest = [record for record in sorted_records if record.article_date == latest_date]
        table_title = latest[0].table_title or f"Bank rates on {latest_date:%Y-%m-%d}"
        latest_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(months))
        latest_ws["A1"] = table_title
        latest_ws["A1"].font = Font(bold=True, size=13, color="1F2937")
        latest_ws["A1"].fill = PatternFill("solid", fgColor="E8F1F8")
        latest_ws["A1"].alignment = Alignment(horizontal="center")
        latest_ws.append(["NGÂN HÀNG"] + [f"{month} THÁNG" for month in months])
        for record in latest:
            latest_ws.append([record.bank] + [record.rates.get(month) for month in months])
        style_header(latest_ws, row=2)
        for row in range(3, latest_ws.max_row + 1):
            for col in range(2, 2 + len(months)):
                latest_ws.cell(row, col).number_format = "0.00"
                latest_ws.cell(row, col).alignment = Alignment(horizontal="right")
        latest_ws.freeze_panes = "B3"
    apply_common_sheet_style(latest_ws)

    articles_ws = wb.create_sheet("Articles")
    articles_ws.append(["date", "status", "rows", "found_on_page", "title", "url", "error"])
    for log in sorted(
        article_logs,
        key=lambda item: (item.article_date or date.min, item.url),
        reverse=True,
    ):
        articles_ws.append(
            [
                log.article_date,
                log.status,
                log.rows,
                log.found_on_page,
                log.title,
                log.url,
                log.error,
            ]
        )
    style_header(articles_ws)
    for row in range(2, articles_ws.max_row + 1):
        articles_ws.cell(row, 1).number_format = "yyyy-mm-dd"
        url_cell = articles_ws.cell(row, 6)
        if url_cell.value:
            url_cell.hyperlink = str(url_cell.value)
            url_cell.style = "Hyperlink"
    add_table(articles_ws, "Articles")
    apply_common_sheet_style(articles_ws)

    wb.save(output_path)


def run_scrape(
    args: argparse.Namespace,
    progress_callback: ProgressCallback | None = None,
) -> dict[str, object]:
    date_from = parse_iso_date(args.date_from)
    date_to = parse_iso_date(args.date_to)
    if date_from and date_to and date_from > date_to:
        raise SystemExit("--date-from must be before --date-to.")

    session = make_session()
    records: list[RateRecord] = []
    article_logs: list[ArticleLog] = []

    if args.legacy:
        emit_progress(progress_callback, "stage", message="Searching legacy news sources...")
        legacy_records, legacy_logs = run_legacy_scrape(session, args, progress_callback)
        records.extend(legacy_records)
        article_logs.extend(legacy_logs)

    if not args.legacy_only:
        if not args.quiet:
            print("Collecting article links...")
        emit_progress(progress_callback, "stage", message="Collecting Vietnamnet article links...")
        links = collect_article_links(session, args, date_from, date_to, progress_callback)
        if not args.quiet:
            print(f"Collected {len(links)} unique article links.")
        emit_progress(
            progress_callback,
            "links_ready",
            total_links=len(links),
            message=f"Collected {len(links)} unique Vietnamnet article links.",
        )

        for idx, link in enumerate(links, start=1):
            try:
                article_records, log = scrape_article(session, link, args)
                records.extend(article_records)
                article_logs.append(log)
                if not args.quiet:
                    print(
                        f"{progress_line(idx, len(links), 'Articles')} | "
                        f"{log.status}: {log.rows} rows - {log.title}"
                    )
                emit_progress(
                    progress_callback,
                    "article",
                    done=idx,
                    total=len(links),
                    status=log.status,
                    rows=log.rows,
                    title=log.title,
                    url=log.url,
                )
            except Exception as exc:  # noqa: BLE001 - continue archive scraping after one bad page.
                article_date = parse_date_from_text(link.title) or parse_date_from_url(link.url)
                log = ArticleLog(
                    article_date=article_date,
                    title=link.title,
                    url=link.url,
                    status="error",
                    error=str(exc),
                    found_on_page=link.found_on_page,
                )
                article_logs.append(log)
                if not args.quiet:
                    if args.verbose:
                        print(
                            f"{progress_line(idx, len(links), 'Articles')} | "
                            f"error: {link.title} - {exc}"
                        )
                    else:
                        print(
                            f"{progress_line(idx, len(links), 'Articles')} | "
                            f"error: {link.title}"
                        )
                emit_progress(
                    progress_callback,
                    "article",
                    done=idx,
                    total=len(links),
                    status="error",
                    rows=0,
                    title=link.title,
                    url=link.url,
                    error=str(exc),
                )

    output_path = Path(args.output)
    run_info = {
        "base_url": args.base_url,
        "pages": args.pages,
        "start_page": args.start_page,
        "date_from": args.date_from,
        "date_to": args.date_to,
    }
    emit_progress(progress_callback, "stage", message="Writing Excel workbook...")
    write_workbook(records, article_logs, output_path, run_info)

    result = {
        "output": str(output_path),
        "articles": len(article_logs),
        "ok_articles": sum(1 for log in article_logs if log.status == "ok"),
        "rate_rows": len(records),
    }
    emit_progress(progress_callback, "finished", **result)
    return result


def main() -> int:
    args = parse_args()
    if args.menu or len(sys.argv) == 1:
        args = apply_menu(args)

    result = run_scrape(args)
    print(json.dumps(result, ensure_ascii=False))
    if not args.quiet:
        print()
        print("Done.")
        print(f"Excel file: {result['output']}")
        print(f"Articles scanned: {result['articles']}")
        print(f"Articles with tables: {result['ok_articles']}")
        print(f"Rate rows exported: {result['rate_rows']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
